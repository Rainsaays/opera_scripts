#/usr/bin/env python3
# -*- coding: utf-8 -*-


import xlsxwriter, pymysql,time,os,openpyxl,chardet
from time import strftime, localtime
from openpyxl import Workbook,load_workbook

from datetime import datetime
from message import message
mysql_host = "192.168.16.31"

mysql_user = "user"
mysql_port = 3306
mysql_passwd = "passwd"
mysql_db = "db"
file_path = "E:\productinfo"

file_name = os.path.join(file_path, ''.join(["productinfo-", strftime("%Y%m%d", localtime()), ".xlsx"]))

commodity_info = '''SELECT 1 from user where user=%s;'''
merchant_info = '''SELECT 1 from user where user=%s;;'''
class data_excel():
    def __init__(self,sql,sheet_name,title_row):
        self.sql = sql
        self.sheet_name = sheet_name
        self.title_row = title_row

###执行sql并写入到excel		
    def sql_exec(self):
        conn = pymysql.connect(host=mysql_host, port=mysql_port, user=mysql_user, password=mysql_passwd, db=mysql_db,
                               charset='utf8')
        cursor = conn.cursor()
        cursor.execute(self.sql,(i))

        row_num = 2

        workbook,worksheet = self.xlsx_open()
        # print(workbook,worksheet)
        for i in cursor.fetchall():
            self.xlsx_writer(worksheet,row_num,i)
            row_num += 1
        self.xlsx_close(workbook)
        cursor.close()

###建立excel
    def xlsx_open(self):
        if os.path.isfile(file_name):
            workbook = load_workbook(file_name)

        else:
            workbook = Workbook()  # 建立文件
            workbook.remove( workbook[ workbook.sheetnames[0] ] )

        if self.sheet_name in workbook.sheetnames:
            workbook.remove( workbook[self.sheet_name])
        worksheet = workbook.create_sheet(self.sheet_name)
        worksheet.append( self.title_row )
        # print(workbook, worksheet)
        return workbook,worksheet
###写入excel
    def xlsx_writer(self,worksheet, row_num, writer_data):
        # print(worksheet, row_num, writer_data)
        col = 1
        for i in writer_data:
            # print( i )
            worksheet.cell(row_num,col,str(i))
            col += 1


###保存excel
    def xlsx_close(self,workbook):
        workbook.save(file_name)


if __name__ == "__main__":
    title_row_commodity = ["123","123"]
    title_row_merchant = ["123","123"]
    commodity_exec=data_excel(commodity_info,"commodity",title_row_commodity)
    commodity_exec.sql_exec()
    merchant_exec=data_excel(merchant_info, "merchant", title_row_merchant)
    merchant_exec.sql_exec()
    message("数据", "", file_name)


